# import os
import ast
import os

import numpy as np
import openpyxl
import torch
from transformers import AutoModel, AutoTokenizer, BertTokenizer, BertModel

from sklearn.metrics.pairwise import cosine_similarity
from vncorenlp import VnCoreNLP

# from get_verb import vncorenlp

# Tải mô hình PhoBERT và tokenizer
model_name = "vinai/phobert-base"
tokenizer_phoBert = AutoTokenizer.from_pretrained(model_name)
model_phoBert = AutoModel.from_pretrained(model_name)

tokenizer = BertTokenizer.from_pretrained('bert-base-multilingual-cased')
model = BertModel.from_pretrained("bert-base-multilingual-cased")
vncorenlp = VnCoreNLP("/home/linhhm/Kmean/VnCoreNLP/VnCoreNLP-1.2.jar")


def tokenize_sentence(text):
    sentences = vncorenlp.tokenize(text)
    data = []
    for sentence in sentences:
        data.append(" ".join(sentence))
    return data[0]


def get_vector_by_phoBert(word, text):
    token = tokenizer_phoBert.encode(text)
    token_word = tokenizer_phoBert.encode(word)[1]
    if token_word not in token:
        return 0
    word_index = token.index(token_word)
    input_ids = torch.tensor([token])
    with torch.no_grad():
        outputs = model_phoBert(input_ids)

    word_embedding = outputs.last_hidden_state[:, word_index, :].tolist()
    return word_embedding[0]


def get_vector_by_baseBert(word, text):
    tokens = tokenizer.tokenize(tokenizer.decode(tokenizer.encode(text)))
    start = 0
    end = 0
    encoded_input = tokenizer(text, return_tensors='pt')
    for tk in tokens:
        if word.startswith(tk.replace("##", "")):
            start = tokens.index(tk)
        if word.endswith(tk.replace("##", "")):
            end = tokens.index(tk)

    outputs = model(**encoded_input)
    target_word_embedding = outputs.last_hidden_state[:, start, :]
    for i in range(start + 1, end + 1):
        target_word_embedding += outputs.last_hidden_state[:, i, :]
    return target_word_embedding.tolist()[0]


def get_vector(words, example):
    vectors = []
    vectors_str = []
    for i in range(len(words)):
        word = str(words[i]).split("_")[0]
        exmpls = example[i].split(";")
        vect = []
        for ex in exmpls:
            emb = get_vector_by_baseBert(word, ex)
            vect.append(emb)
        vectors.append(average_vecto(vect))
        vectors_str.append(str(average_vecto(vect)))
    return vectors, vectors_str


def average_vecto(data):
    return np.mean(np.array(data), axis=0)


def get_vecto_cluster(V, example):
    averages = []
    for i in range(len(V)):
        v = V[i][1:]
        examp = example[i][1:]
        vecto, _ = get_vector(v, examp)
        averages.append(list(average_vecto(vecto)))
    return averages


def get_vecto_english(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb["Sheet1"]
    max_row = sheet.max_row

    clusters = {}
    for row in range(1, max_row + 1):
        cluster = sheet[f"A{row}"].value
        vecto = sheet[f"D{row}"].value
        vecto_list = ast.literal_eval(str(vecto))
        if cluster.strip() not in clusters.keys():
            clusters[cluster.strip()] = [vecto_list]
        else:
            clusters[cluster.strip()].append(vecto_list)
        # if cluster.strip() == "Put 9.1":
        #     print(vecto_list)
    cluster_key = list(clusters.keys())
    vector_cluster = []
    for i in list(clusters.values()):
        vector_cluster.append(list(average_vecto(i)))
    return cluster_key, vector_cluster


def output_Excel(matrix, file_name):
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet()
    for row in matrix:
        sheet.append(row)

    wb.save(file_name)


def cousin_sim(vector_a, vector_b):
    print(len(vector_a), len(vector_b))
    cosine_sim = cosine_similarity(np.asarray(vector_a).reshape(1, -1), np.asarray(vector_b).reshape(1, -1))[0][0]
    return round(cosine_sim, 4)


def get_matrix(cluster_vn, vecto_vn, cluster_en, vecto_en):
    matrix = []
    row1 = [""] + list(cluster_en)
    matrix.append(row1)
    for i in range(len(vecto_vn)):
        row_i = [cluster_vn[i]]
        for en in vecto_en:
            row_i.append(cousin_sim(vecto_vn[i], en))
        matrix.append(row_i)
    return matrix


def get_10_max_idx(arr):
    sorted_indices = np.argsort(np.array(arr))[::-1]
    top_10_indices = sorted_indices[1:11]
    return top_10_indices


def get_10_max(cluster_vn, vecto_vn, cluster_en, vecto_en):
    matrix = []
    # row1 = ["", ""]
    # matrix.append(row1)
    idx = {}
    cosin = {}
    for i in range(len(vecto_vn)):
        # row = {}
        row_i = []
        for en in vecto_en:
            row_i.append(cousin_sim(vecto_vn[i], en))
        idx_max = get_10_max_idx(row_i)
        cosin[i] = row_i
        idx[i] = idx_max
    id_arr = []
    for i in idx.values():
        for j in i:
            id_arr.append(j)
    set_idx = set(id_arr)
    row1 = [""]
    for i in list(set_idx):
        row1.append(cluster_en[i])
    matrix.append(row1)
    for i in range(len(vecto_vn)):
        row_i = [""] * len(row1)
        row_i[0] = [cluster_vn[i]]
        for j in idx[i]:
            idx_in_key = int(list(set_idx).index(j))
            row_i[idx_in_key] = cosin[j]
        matrix.append(row_i)

    print(len(matrix))
    print(len(matrix[0]), len(matrix[1]))
    return matrix


# verb, examples, vn_cluster = get_data("Data/Output/verb_examples/10")
# average_cluster_vn = get_vecto_cluster(verb, examples)
# en_cluster, vecto_cluster = get_vecto_english("Data/verb_english.xlsx")
# # matrix_en_vn = get_matrix(vn_cluster, average_cluster_vn, en_cluster, vecto_cluster)
#
# matrix_en_vn = get_10_max(vn_cluster, average_cluster_vn, en_cluster, vecto_cluster)
# output_Excel(matrix_en_vn, "Data/cluster_by_defin_vn_en_10_max.xlsx")
